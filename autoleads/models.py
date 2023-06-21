from django.db import models
from django.forms.models import model_to_dict
from django.db import transaction
from django.utils.translation import gettext_lazy as _
from django.http import JsonResponse
import re
import os
from pathlib import Path
from django.conf import settings
from openpyxl import load_workbook
from unidecode import unidecode
from django.utils import timezone
import json


class App(models.Model):
    name = models.CharField(_('App name'), max_length=100, unique=True)
    created_by = models.EmailField(_('Created by'), default='')
    created_time = models.DateTimeField(_('Date created'), default=timezone.now)
    edited_by = models.EmailField(_('Edited by'), default='')
    edited_time = models.DateTimeField(_('Date edited'), default=timezone.now)
    white_list = models.TextField(_('White list'), default='amazon, ebay')
    black_list = models.TextField(_('Black list'), default='')
    loop_time = models.CharField(_('Loop time'), max_length=10, default='60')
    amazon_base_url = models.URLField(_('Amazon base URL'), max_length=200, default='https://www.amazon.com/dp/')
    compare_value = models.CharField(_('Compare value'), max_length=10, default='35')
    auto_restart_value = models.BooleanField(_('Auto restart'), default=True)
    force_restart_count = models.CharField(_('Force restart count'), max_length=10, default='0')
    auto_restart_count = models.CharField(_('Auto restart count'), max_length=10, default='0')
        
    class Meta:
        app_label = 'autoleads'
    
    def __str__(self):
        return self.name

    @classmethod
    @transaction.atomic
    def create_or_update_multiple(cls, user, bulk_data):
        try:
            for data in bulk_data:
                excluded_keys = {'amazon_products'}
                defaults = {k: v for k, v in data.items() if k not in excluded_keys}
                defaults.update({"edited_by": user.email, "edited_time": timezone.now()})
                app_item, created = cls.objects.update_or_create(
                    name=data.get('name'),
                    defaults=defaults
                )

                amazon_product_data_list = data.get('amazon_products', [])
                product_ASINs = [amazon_product_data.get('ASIN') for amazon_product_data in amazon_product_data_list]
                AmazonProduct.objects.filter(app=app_item).exclude(ASIN__in=product_ASINs).delete()
                for amazon_product_data in amazon_product_data_list:
                    AmazonProduct.objects.update_or_create(
                        app=app_item,
                        ASIN=amazon_product_data.get('ASIN'),
                        defaults=amazon_product_data
                    )

                return {'success': True, 'message': 'App data updated/created successfully.'}

        except Exception as e:
            return {'success': False, 'error': str(e)}

    @classmethod
    def get_all(cls, user):
        try:
            app_items = cls.objects.all().prefetch_related('amazon_products')
            if app_items.exists():
                data = []
                for app in app_items:
                    app_data = model_to_dict(app)
                    app_data.pop('id', None)
                    app_data.pop('created_by', None)
                    app_data.pop('created_time', None)
                    app_data.pop('edited_by', None)
                    app_data.pop('edited_time', None)
                    app_data.pop('force_restart_count', None)
                    app_data.pop('auto_restart_count', None)

                    app_data['amazon_products'] = []
                    for product in app.amazon_products.all():
                        product_data = model_to_dict(product)
                        product_data.pop('id', None)
                        product_data.pop('app', None)
                        app_data['amazon_products'].append(product_data)

                    data.append(app_data)
                return {'success': True, 'data': data}

            default_app = cls.objects.create(name="Default App", created_by=user, edited_by=user)

            default_app_data = list(cls.objects.filter(id=default_app.id).prefetch_related('amazon_products').values())
            return {'success': True, 'data': default_app_data}

        except Exception as e:
            return {'success': False, 'error': str(e)}

class AmazonProduct(models.Model):
    app = models.ForeignKey(App, on_delete=models.CASCADE, related_name='amazon_products')
    ASIN = models.CharField(_('ASIN'), max_length=10, unique=True)
    status = models.BooleanField(_('Status'), default=True)

    class Meta:
        app_label = 'autoleads'

class AppService:
    @staticmethod
    def set_all(user, apps):
        if not apps:
            return {'success': False, 'error': 'Data is not available'}

        response = App.create_or_update_multiple(user, apps)
        if response['success']:
            APP_DIR = Path(__file__).resolve().parent
            with open(os.path.join(APP_DIR, 'app/apps.json'), 'w') as f:
                json.dump(apps, f)
        return response
        
    @staticmethod
    def get_all(user):
        apps = App.get_all(user=user)
        return {'success': True, 'data': apps}

class AuxiliaryService:
    @staticmethod
    def upload_product_files(user, files):
        if not files:
            return JsonResponse({'success': False, 'error': 'Files not found'})
        
        def validate_asin(asin):
            asin_regex = r'^[a-zA-Z0-9]{10}$'
            return re.match(asin_regex, str(asin)) is not None
        
        def remove_file(file_path):
            if os.path.exists(file_path):
                os.remove(file_path)

        user_folder_dir = os.path.join(settings.MEDIA_ROOT, str(user))
        if not os.path.exists(user_folder_dir):
            os.makedirs(user_folder_dir)

        file_errors = {'success': False, 'error': ''}
        file_paths = []
        for file in files:
            file_path = os.path.join(user_folder_dir, f'{file.name}')
            try:
                with open(file_path, 'wb') as destination:
                    for chunk in file.chunks():
                        destination.write(chunk)
                file_paths.append(file_path)
            except Exception as e:
                file_errors['error'] += f'{e}\t'
        
        if not file_paths:
            return JsonResponse(file_errors)
                
        products = []
        total_count = 0
        validate_count = 0
        MAX_PRODUCT_COUNT = 1000
        error_responses = ''

        try:
            for file in files:
                if file.name.endswith('.xlsx'):
                    wb = load_workbook(file_path)
                    ws = wb.active
                    header_row = next(ws.iter_rows(min_row=0, max_row=3, values_only=True))
                    asin_column = None
                    for i, header in enumerate(header_row):
                        _header = unidecode(str(header)).strip().upper().replace(' ', '_')
                        if _header == 'none':
                            continue
                        if asin_column is None and ('ASIN' in _header):
                            asin_column = i + 1

                    if asin_column is None:
                        for row in ws.iter_rows(min_row=0, max_row=3, values_only=True):
                            for i, cell_value in enumerate(row):
                                if validate_asin(cell_value):
                                    asin_column = i + 1
                                    break

                    if asin_column is None:
                        remove_file(file_path)
                        error_responses += f'In {file.name} asin column not found\t'

                    if asin_column is None:
                        continue

                    seen_asins = set()
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if len(seen_asins) >= MAX_PRODUCT_COUNT:
                            break
                        asin = row[asin_column - 1]
                        if asin is None or asin in seen_asins:
                            continue
                        seen_asins.add(asin)
                        total_count += 1
                        if asin and validate_asin(asin):
                            asin_data = {'ASIN': asin}
                            asin_data['status'] = True
                            products.append(asin_data)
                            validate_count += 1
                else:
                    remove_file(file_path)
                    return {'success': False, 'error': 'Unsupported file type'}
        except Exception as e:
            remove_file(file_path)
            return {'success': False, 'error': f'An error occurred: {str(e)}'}
        finally:
            remove_file(file_path)
        return {'success': True, 'products': products, 'total_count': total_count, 'validate_count': validate_count, 'unknown_count': total_count - validate_count}
