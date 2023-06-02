import os
import re
from openpyxl import load_workbook
from unidecode import unidecode

from django.db import models
from django.conf import settings

from pprint import pprint
from ElasticEmail import ApiClient, Configuration, ApiException
from ElasticEmail.api.files_api import FilesApi
from ElasticEmail.model.file_payload import FilePayload
import base64
import mimetypes

from ElasticEmail.api.emails_api import EmailsApi
from ElasticEmail.model.email_content import EmailContent
from ElasticEmail.model.body_part import BodyPart
from ElasticEmail.model.body_content_type import BodyContentType
from ElasticEmail.model.email_recipient import EmailRecipient
from ElasticEmail.model.email_message_data import EmailMessageData
from datetime import datetime

configuration = Configuration(
    host = "https://api.elasticemail.com/v4"
)
configuration.api_key['apikey'] = os.getenv('ELASTIC_EMAIL_API_KEY')


###
# HELPER FUNCTIONS
###
def remove_file(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)

def create_fields_dict(item):
    fields_dict = {}
    for field_key, field_value in item.items():
        if field_key not in ["email", "status"]:
            fields_dict[field_key] = field_value
    return fields_dict
###
##################
###

class UserTemplates(models.Model):
    user = models.ForeignKey(settings.AUTH_USER_MODEL, on_delete=models.CASCADE)
    name = models.CharField(max_length=255)
    template = models.TextField()

    @classmethod
    def save_template_content(cls, user, name, template):
        user_template = cls.objects.filter(user=user, name=name)
        if user_template:
            user_template.delete()

        user_template = cls(user=user, name=name, template=template)
        user_template.save()
        return {'success': True}
    
    @classmethod
    def get_template_content(cls, user):
        return {'success': True, 'templates': list(UserTemplates.objects.filter(user=user).values().distinct())}
    
    @classmethod
    def delete_template_content(cls, user, name):
        user_template = cls.objects.filter(user=user, name=name)
        if user_template:
            user_template.delete()
            return {'success': True, 'name': name}
        return {'success': False, 'error': f'{name} template not found'}


class UserEmails(models.Model):
    user = models.ForeignKey(settings.AUTH_USER_MODEL, on_delete=models.CASCADE)
    name = models.CharField(max_length=255)
    emails = models.JSONField()

    @classmethod
    def process_email_file(cls, user, file):
        def validate_email(email):
            email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            return re.match(email_regex, str(email)) is not None

        user_folder_dir = os.path.join(settings.MEDIA_ROOT, str(user))
        if not os.path.exists(user_folder_dir):
            os.makedirs(user_folder_dir)
        file_path = os.path.join(user_folder_dir, f'{file.name}')
        with open(file_path, 'wb') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        emails = []
        total_count = 0
        validate_count = 0
        MAX_EMAIL_COUNT = 100

        try:
            if file.name.endswith('.xlsx'):
                wb = load_workbook(file_path)
                ws = wb.active
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                email_column = None
                fields = {}
                for i, header in enumerate(header_row):
                    _header = unidecode(str(header)).strip().lower().replace(' ', '_')
                    if _header == 'none':
                        continue
                    if email_column is None and _header == 'email':
                        email_column = i + 1
                    else:
                        fields[_header] = i + 1

                if email_column is None:
                    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                        for i, cell_value in enumerate(row):
                            if validate_email(cell_value):
                                email_column = i + 1
                                _header = unidecode(str(header)).strip().lower().replace(' ', '_')
                                del fields[_header]
                                break

                if email_column is None:
                    remove_file(file_path)
                    return {'success': False, 'error': 'Email column not found'}

                seen_emails = set()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(seen_emails) >= MAX_EMAIL_COUNT:
                        break
                    email = row[email_column - 1]
                    if email is None or email in seen_emails:
                        continue
                    seen_emails.add(email)
                    total_count += 1
                    if email and validate_email(email):
                        email_data = {'email': email}
                        for field, col_index in fields.items():
                            email_data[field] = row[col_index - 1]
                        email_data['status'] = True
                        emails.append(email_data)
                        validate_count += 1
            else:
                remove_file(file_path)
                return {'success': False, 'error': 'Unsupported file type'}
        except Exception as e:
            remove_file(file_path)
            return {'success': False, 'error': f'An error occurred: {str(e)}'}
        finally:
            remove_file(file_path)
        return {'success': True, 'emails': emails, 'total_count': total_count, 'validate_count': validate_count, 'unknown_count': total_count - validate_count}
    
    @classmethod
    def get_all_email_names(cls, user):
        return {'success': True, 'names': list(cls.objects.filter(user=user).values_list('name', flat=True).distinct())}
    
    @classmethod
    def save_emails(cls, user, name, emails):
        user_emails = cls.objects.filter(user=user, name=name)
        if user_emails:
            user_emails.delete()
        if len(emails) != 0:
            user_emails = cls(user=user, name=name, emails=emails)
            user_emails.save()
        return {'success': True}
    
    @classmethod
    def get_emails(cls, user, name):
        emails = cls.objects.filter(user=user, name=name).values_list('emails', flat=True).distinct()
        return {'success': True, 'emails': list(emails)} if len(emails) else {'success': False, 'error': 'Failed to fetch emails'}
    
    @staticmethod
    def upload_file_to_elasticemail(file_path):
        with ApiClient(configuration) as api_client:
            # api_client.rest_client.pool_manager.connection_pool_kw['cert_reqs'] = ssl.CERT_NONE
            api_instance = FilesApi(api_client)

            with open(file_path, 'rb') as file:
                file_content = file.read()
                binary_content = base64.b64encode(file_content).decode('utf-8')

            content_type = mimetypes.guess_type(file_path)[0] or 'application/octet-stream'
            file_name = os.path.basename(file_path)
            file_name_without_ext, file_ext = os.path.splitext(file_name)
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            unique_file_name = f"{file_name_without_ext}_{timestamp}{file_ext}"

            file_payload = FilePayload(
                binary_content=binary_content,
                name=unique_file_name,
                content_type=content_type,
            )
            expires_after_days = 1

            try:
                apply_result = api_instance.files_post(file_payload, expires_after_days=expires_after_days, async_req=True)
                response = apply_result.get()
                return {'success': True, 'response': {'FileID': response.FileID, 'content_type': response.content_type, 'size': response.size}, 'unique_name': response.file_name}
            except ApiException as e:
                error_message = f"Error: {e.status} - {e.reason}"
                return {'success': False, 'error': error_message}

    @classmethod
    def upload_attachments(cls, user, files):
        tasks = []
        success_responses = []
        error_responses = []
        user_folder_dir = os.path.join(settings.MEDIA_ROOT, str(user))

        if not os.path.exists(user_folder_dir):
            os.makedirs(user_folder_dir)
        for file in files:
            try:
                file_path = os.path.join(user_folder_dir, f'{file.name}')
                with open(file_path, 'wb') as destination:
                    for chunk in file.chunks():
                        destination.write(chunk)
                task = cls.upload_file_to_elasticemail(file_path)
                tasks.append(task)
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as e:
                error_responses.append({'name': file.name, 'error': str(e)})

        for i, task in enumerate(tasks):
            result = task
            file = files[i]
            if isinstance(result, dict) and result.get('success'):
                success_responses.append({
                    'name': file.name,
                    'unique_name': result['unique_name']
                })
            else:
                error_responses.append({'name': file.name, 'error': result.get('error')})

        return {'success': True, 'success_responses': success_responses, 'error_responses': error_responses}
    
    @classmethod
    def send_bulk_emails_via_elasticemail(cls, email_data):
        with ApiClient(configuration) as api_client:
            # api_client.rest_client.pool_manager.connection_pool_kw['cert_reqs'] = ssl.CERT_NONE
            api_instance = EmailsApi(api_client)
            email_message_data = EmailMessageData(
                recipients=[
                    EmailRecipient(email=item['email'], fields=create_fields_dict(item)) for item in email_data['email_list']
                ],
                content=EmailContent(
                    body=[
                        BodyPart(
                            content_type=BodyContentType("PlainText"),
                            content=f"{email_data['message']}",
                            charset="utf-8",
                        ),
                    ],
                    _from=email_data['from'],
                    reply_to=email_data['replyTo'],
                    subject=email_data['subject'],
                    attach_files=email_data['attachment_list'],
                ),
            )
            
            try:
                apply_result = api_instance.emails_post(email_message_data, async_req=True)
                response = apply_result.get()
                return {'success': True, 'response': {'message_id': response.message_id, 'transaction_id': response.transaction_id}}
            except ApiException as e:
                error_message = f"Error: {e.status} - {e.reason}"
                return {'success': False, 'error': error_message}