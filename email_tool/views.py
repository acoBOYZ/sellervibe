from django.shortcuts import render
import os
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from django.http import JsonResponse
import re
from django.contrib.auth.decorators import login_required
from .models import UserEmails
import json
from datetime import datetime
from unidecode import unidecode

import ssl
from pprint import pprint
from ElasticEmail import ApiClient, Configuration, ApiException
from ElasticEmail.api.files_api import FilesApi
from ElasticEmail.model.file_payload import FilePayload
import base64
import mimetypes

from ElasticEmail.api.emails_api import EmailsApi
from ElasticEmail.model.email_content import EmailContent
from ElasticEmail.model.body_part import BodyPart
from ElasticEmail.model.body_content_type import BodyContentType
from ElasticEmail.model.email_recipient import EmailRecipient
from ElasticEmail.model.email_message_data import EmailMessageData

configuration = Configuration(
    host = "https://api.elasticemail.com/v4"
)
configuration.api_key['apikey'] = os.getenv('ELASTIC_EMAIL_API_KEY')

def mail(request):
    return render(request, 'mail/index.html')

def contacts(request):
    return render(request, 'contacts/index.html')

def templates(request):
    return render(request, 'templates/index.html')

def signatures(request):
    return render(request, 'signatures/index.html')

def send(request):
    return render(request, 'send/index.html')


@csrf_exempt
def upload_file(request):
    if request.method == 'POST':
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'error': 'Authentication required'})
        file = request.FILES.get('file')
        if file:
            def validate_email(email):
                email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                return re.match(email_regex, str(email)) is not None
            
            file_path = os.path.join(f'{file.name}')
            with open(file_path, 'wb') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
            emails = []
            total_count = 0
            validate_count = 0
            MAX_EMAIL_COUNT = 100
            try:
                if file.name.endswith('.xlsx'):
                    wb = load_workbook(file_path)
                    ws = wb.active
                    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    email_column = None
                    fields = {}
                    for i, header in enumerate(header_row):
                        _header = unidecode(str(header)).strip().lower().replace(' ', '_')
                        if _header == 'none':
                            continue
                        if email_column is None and _header == 'email':
                            email_column = i + 1
                        else:
                            fields[_header] = i + 1

                    if email_column is None:
                        for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                            for i, cell_value in enumerate(row):
                                if validate_email(cell_value):
                                    email_column = i + 1
                                    _header = unidecode(str(header)).strip().lower().replace(' ', '_')
                                    del fields[_header]
                                    break

                    if email_column is None:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                        return JsonResponse({'success': False, 'error': 'Email column not found'})

                    seen_emails = set()
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if len(seen_emails) >= MAX_EMAIL_COUNT:
                            break
                        email = row[email_column - 1]
                        if email is None or email in seen_emails:
                            continue
                        seen_emails.add(email)
                        total_count += 1
                        if email and validate_email(email):
                            email_data = {'email': email}
                            for field, col_index in fields.items():
                                email_data[field] = row[col_index - 1]
                            email_data['status'] = True
                            emails.append(email_data)
                            validate_count += 1
                else:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                    return JsonResponse({'success': False, 'error': 'Unsupported file type'})
            except Exception as e:
                if os.path.exists(file_path):
                    os.remove(file_path)
                return JsonResponse({'success': False, 'error': f'An error occurred: {str(e)}'})
            finally:
                if os.path.exists(file_path):
                    os.remove(file_path)
            return JsonResponse({'success': True, 'emails': emails, 'total_count': total_count, 'validate_count': validate_count, 'unknown_count': total_count - validate_count})
    return JsonResponse({'success': False})

def get_email_names(request):
    if not request.user.is_authenticated:
        return JsonResponse({'names': []})
    email_file_names = UserEmails.objects.filter(user=request.user).values_list('name', flat=True).distinct()
    return JsonResponse({'names': list(email_file_names)})

@csrf_exempt
def save_emails(request):
    if request.method == 'POST' and request.user.is_authenticated:
        emails = json.loads(request.POST.get('emails'))
        delete_all = len(emails) == 0
        name = request.POST.get('name')
        user_emails = UserEmails.objects.filter(user=request.user, name=name)
        user_emails.delete()
        if not delete_all:
            user_emails = UserEmails(user=request.user, name=name, emails=emails)
            user_emails.save()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

def get_emails(request):
    if not request.user.is_authenticated:
        return JsonResponse({'success': False})
    name = request.GET.get('name') or request.POST.get('name')
    queryset = UserEmails.objects.filter(user=request.user)
    if name:
        queryset = queryset.filter(name=name)
    emails = queryset.values_list('emails', flat=True).distinct()
    return JsonResponse({'success': True, 'emails': list(emails)}, safe=False)


def upload_file_to_elasticemail(file_path):
    with ApiClient(configuration) as api_client:
        api_client.rest_client.pool_manager.connection_pool_kw['cert_reqs'] = ssl.CERT_NONE
        api_instance = FilesApi(api_client)

        with open(file_path, 'rb') as file:
            file_content = file.read()
            binary_content = base64.b64encode(file_content).decode('utf-8')

        content_type = mimetypes.guess_type(file_path)[0] or 'application/octet-stream'
        file_name = os.path.basename(file_path)
        file_name_without_ext, file_ext = os.path.splitext(file_name)
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        unique_file_name = f"{file_name_without_ext}_{timestamp}{file_ext}"

        file_payload = FilePayload(
            binary_content=binary_content,
            name=unique_file_name,
            content_type=content_type,
        )
        expires_after_days = 1

        try:
            apply_result = api_instance.files_post(file_payload, expires_after_days=expires_after_days, async_req=True)
            response = apply_result.get()
            return {'success': True, 'response': {'FileID': response.FileID, 'content_type': response.content_type, 'size': response.size}, 'unique_name': response.file_name}
        except ApiException as e:
            error_message = f"Error: {e.status} - {e.reason}"
            print(f"Error details: {e.body}")
            return {'success': False, 'error': error_message}

@csrf_exempt
def upload_attachments(request):
    if request.method == 'POST' and request.user.is_authenticated:
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'error': 'Authentication required'})
        tasks = []
        for file in request.FILES.getlist('attachment'):
            file_path = os.path.join(f'{file.name}')
            with open(file_path, 'wb') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
            task = upload_file_to_elasticemail(file_path)
            tasks.append(task)
            if os.path.exists(file_path):
                os.remove(file_path)

        success_responses = []
        error_responses = []

        for i, task in enumerate(tasks):
            result = task
            file = request.FILES.getlist('attachment')[i]
            if isinstance(result, dict) and result.get('success'):
                success_responses.append({
                    'success': True,
                    'name': file.name,
                    'unique_name': result['unique_name']
                })
            else:
                error_responses.append({'success': False, 'name': file.name, 'error': result.get('error')})

        return JsonResponse({'success_responses': success_responses, 'error_responses': error_responses})
    else:
        return JsonResponse({'success': False})


def create_fields_dict(item):
    fields_dict = {}
    for field_key, field_value in item.items():
        if field_key not in ["email", "status"]:
            fields_dict[field_key] = field_value
    return fields_dict

def send_bulk_emails_via_elasticemail(email_data):
    with ApiClient(configuration) as api_client:
        api_client.rest_client.pool_manager.connection_pool_kw['cert_reqs'] = ssl.CERT_NONE
        api_instance = EmailsApi(api_client)
        email_message_data = EmailMessageData(
            recipients=[
                EmailRecipient(email=item['email'], fields=create_fields_dict(item)) for item in email_data['email_list']
            ],
            content=EmailContent(
                body=[
                    BodyPart(
                        content_type=BodyContentType("PlainText"),
                        content=f"{email_data['message']}",
                        charset="utf-8",
                    ),
                ],
                _from=email_data['from'],
                reply_to=email_data['replyTo'],
                subject=email_data['subject'],
                attach_files=email_data['attachment_list'],
            ),
        )
        
        try:
            apply_result = api_instance.emails_post(email_message_data, async_req=True)
            response = apply_result.get()
            return {'success': True, 'response': {'message_id': response.message_id, 'transaction_id': response.transaction_id}}
        except ApiException as e:
            error_message = f"Error: {e.status} - {e.reason}"
            return {'success': False, 'error': error_message}

@csrf_exempt
def send_bulk_emails(request):
    if request.method == 'POST':
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'error': 'Authentication required'})
        
        attachment_list = []
        email_list = []
        cc_list = []

        for attachment in request.POST.getlist('attachment'):
            attachment_list.append(attachment)
        for email in request.POST.getlist('email'):
            _e = json.loads(email)
            if _e['status']:
                email_list.append(_e)
        if len(email_list) == 0:
            return JsonResponse({'success': False, 'error': 'email list is empty'})
        text_from = request.POST.get('from')
        text_replyTo = request.POST.get('replyTo')
        for cc in request.POST.getlist('cc'):
            cc_list.append(cc)
        text_subject = request.POST.get('subject')
        text_message = request.POST.get('message')

        email_data = {
            'attachment_list': attachment_list,
            'email_list': email_list,
            'cc_list': cc_list,
            'from': text_from,
            'replyTo': text_replyTo,
            'subject': text_subject,
            'message': text_message
        }

        response = send_bulk_emails_via_elasticemail(email_data)
        
        return JsonResponse(response)
    else:
        return JsonResponse({'success': False, 'error': 'method is not valid'})
